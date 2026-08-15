"""把兩份獨立蒐集的展商資料合併成 import_exhibitors.py 吃的 CSV。

使用方式：
    python3 scripts/merge_exhibitor_sources.py 展商.xlsx 抓到的.csv -o merged.csv

兩份來源（2026-08-10）：
- xlsx：GPT 整理的完整名單，欄位最齊（881 家全部有簡中名、展位、Logo、
        地址，另有公司網站 723、產品目錄 PDF 542+229）
- csv ：本專案 scrape_exhibitor_list.js 從官方網頁獨立抓的，欄位較少，
        但是**獨立來源**，拿來交叉驗證 xlsx 有沒有被竄改或憑空生成

交叉驗證結果（合併前實測）：兩邊 881 個展商 ID 完全一致、展會分類
881/881 相同、展位 877/881 相同（那 4 筆差異是 csv 這邊「4G102-8」這種
再細分編號的正規化 bug，已修）。兩份獨立來源互相印證，可以放心採用。

合併原則：
- 以 xlsx 為主要來源（欄位齊全），csv 只在 xlsx 該欄位是空的時候補位
- 兩邊衝突時保留 xlsx 的值，但一律列進報表讓人看得到，不默默選一邊
- 官方展商 ID（xlsx 的「展商ID」＝ csv 的 ex_id）當合併鍵，不用公司名比對
  ——名稱寫法差異在先前的匯入已經吃過虧（Co., Ltd. vs CO LTD）
"""
import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("需要 openpyxl：pip install openpyxl")

# import_exhibitors.py 吃的欄位
OUT_COLS = ["name_zh", "name_en", "booth_no", "hall", "country", "category",
            "description", "website", "directory_url", "pdfs"]


def normalize_booth(raw):
    """2F310 → N2-F310；4G102-8 → N4-G102-8（同一攤位再細分的編號要保留）。

    跟 scrape_exhibitor_list.js 的 normalizeBooth() 是同一套規則，兩邊要
    一起改。認不出來的格式原樣保留（實際資料裡有 GP、Sponser 這種非攤位
    字串），不要因為配不上格式就默默清成空字串。
    """
    s = str(raw or "").strip()
    if not s:
        return ""
    if re.match(r"^[NWE]\d-", s, re.I):
        return s.upper()
    m = re.match(r"^(\d)\s*([A-Za-z])\s*(\d{2,4})(-\d+)?$", s)
    if m:
        return f"N{m.group(1)}-{m.group(2).upper()}{m.group(3)}{m.group(4) or ''}"
    return s


def hall_from_booth(booth):
    """N2-F310 → N2。認不出來就留空，不亂猜。"""
    m = re.match(r"^([NWE]\d)-", booth or "")
    return m.group(1) if m else ""


def parse_categories(raw):
    """官方分類文字 → 既有的 cat-XX / cat-08-X。

    跟 scrape_exhibitor_list.js 的 parseCategories() 同一套規則。
    主分類寫成 "2. Metallic…"（數字＋點＋空格），8.x 子分類寫成
    "8.2Sensing…"（數字後直接接字母，沒有空格也沒有第二個點）。
    """
    text = str(raw or "")
    out, sub_main_seen = [], set()
    for m in re.finditer(r"(\d{1,2})\.(\d)(?!\d)", text):
        out.append(f"cat-{m.group(1).zfill(2)}-{m.group(2)}")
        sub_main_seen.add(m.group(1))
    for m in re.finditer(r"(?:^|[^\d.])(\d{1,2})\.(?!\d)", text):
        if m.group(1) in sub_main_seen:
            continue
        out.append(f"cat-{m.group(1).zfill(2)}")
    seen, uniq = set(), []
    for c in out:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    return uniq


def load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["展商名單"]
    hdr = [c.value for c in ws[1]]
    rows = {}
    for r in ws.iter_rows(min_row=2):
        d = dict(zip(hdr, [c.value for c in r]))
        ex_id = str(d.get("展商ID") or "").strip()
        if ex_id:
            rows[ex_id] = d
    return rows


def load_csv(path):
    rows = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            ex_id = (r.get("ex_id") or "").strip()
            if ex_id:
                rows[ex_id] = r
    return rows


def clean(v):
    s = str(v or "").strip()
    return "" if s.lower() in ("none", "nan", "/") else s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path", type=Path)
    ap.add_argument("csv_path", type=Path)
    ap.add_argument("-o", "--out", type=Path, required=True)
    args = ap.parse_args()

    xl = load_xlsx(args.xlsx_path)
    cs = load_csv(args.csv_path)

    only_xl = set(xl) - set(cs)
    only_cs = set(cs) - set(xl)
    both = set(xl) & set(cs)
    print(f"xlsx {len(xl)} 家｜csv {len(cs)} 家｜兩邊都有 {len(both)} 家")
    if only_xl:
        print(f"  只在 xlsx：{len(only_xl)} 家（仍會收進來，但少了獨立來源佐證）")
    if only_cs:
        print(f"  只在 csv：{len(only_cs)} 家")

    conflicts = {"booth": [], "category": [], "country": []}
    out_rows = []
    for ex_id in sorted(set(xl) | set(cs), key=lambda x: int(x) if x.isdigit() else 0):
        x = xl.get(ex_id, {})
        c = cs.get(ex_id, {})

        booth_x = normalize_booth(x.get("展位"))
        booth_c = clean(c.get("booth_no"))
        if booth_x and booth_c and booth_x != booth_c:
            conflicts["booth"].append((ex_id, booth_x, booth_c))
        booth = booth_x or booth_c

        cats_x = parse_categories(x.get("展會分類"))
        cats_c = [t for t in clean(c.get("categories_all")).split() if t]
        if cats_x and cats_c and cats_x[0] != cats_c[0]:
            conflicts["category"].append((ex_id, cats_x[0], cats_c[0]))
        category = (cats_x or cats_c or [""])[0]

        country_x = clean(x.get("國家／地區"))
        country_c = clean(c.get("country"))
        if country_x and country_c and country_x.upper() != country_c.upper():
            conflicts["country"].append((ex_id, country_x, country_c))
        country = country_x or country_c

        # 型錄：中文版與英文版各一個欄位，兩個都收（匯入端用 ; 分隔多值）
        pdfs = [clean(x.get("產品目錄PDF（主要／中文）")), clean(x.get("產品目錄PDF（英文）"))]
        pdfs = [p for p in pdfs if p.startswith("http")]

        website = clean(x.get("公司網站"))
        if website and not website.startswith("http"):
            website = "https://" + website  # 有些欄位只填網域，補協定才點得開

        out_rows.append({
            "name_zh": clean(x.get("簡體中文名稱")) or clean(c.get("name_zh")),
            "name_en": clean(x.get("英文名稱")) or clean(c.get("name_en")),
            "booth_no": booth,
            "hall": hall_from_booth(booth),
            "country": country,
            "category": category,
            "description": clean(x.get("公司介紹（簡中）")) or clean(x.get("公司介紹（英文）")) or clean(c.get("description")),
            "website": website,
            "directory_url": clean(x.get("展商詳細頁")) or clean(c.get("directory_url")),
            "pdfs": ";".join(pdfs),
        })

    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS)
        w.writeheader()
        w.writerows(out_rows)

    print(f"\n已輸出 {len(out_rows)} 家 → {args.out}")
    for f in OUT_COLS:
        n = sum(1 for r in out_rows if r[f])
        print(f"  {f}: {n}")

    print("\n=== 兩份來源不一致的欄位（保留 xlsx 的值，列出來讓人核對）===")
    for kind, items in conflicts.items():
        print(f"  {kind}: {len(items)} 筆")
        for ex_id, a, b in items[:5]:
            print(f"     id={ex_id}  xlsx={a!r}  csv={b!r}")


if __name__ == "__main__":
    main()
